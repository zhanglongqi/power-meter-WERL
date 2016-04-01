#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
longqi 29/Jan/16 15:24

"""
from pymodbus.client.sync import ModbusTcpClient as ModbusClient
from time import time, localtime, strftime
from logging import getLogger
from random import randint
from openpyxl import Workbook
import os

from meter import SchneiderPM710, MeterData
from utils import FloatData

debug = True


class Panel():
    def __init__(self, ip, name):
        self.client = ModbusClient(ip)

        self.AC_Meter_0 = SchneiderPM710(unit_id=2)
        self.AC_Meter_1 = SchneiderPM710(unit_id=3)

        self.name = name
        self.logger = getLogger(self.name)

        self.book = Workbook()

        self.sheet_for_AC_Meter_0 = self.book.active
        self.sheet_for_AC_Meter_0.title = 'AC_Meter_0'

        self.sheet_for_AC_Meter_1 = self.book.create_sheet(title='AC_Meter_1')
        self.dir = os.path.join(os.path.dirname(os.path.realpath(__file__)),'..','db')
        if not os.path.exists(self.dir):
            os.makedirs(self.dir)

    def read_and_parse_from_ModbusTCP(self):
        for i in range(0, 2):

            data = MeterData()
            data['time'] = time()

            if i == 0:
                res = self.client.read_holding_registers(999, 22, unit=self.AC_Meter_0.unit_id) if not debug else None
            elif i == 1:
                res = self.client.read_holding_registers(999, 22, unit=self.AC_Meter_1.unit_id) if not debug else None

            real_energy = FloatData()
            real_energy.shorts.s0 = res.registers[1] if not debug else -1
            real_energy.shorts.s1 = res.registers[0] if not debug else -1
            data['real_energy'] = real_energy.float if not debug else randint(1, 100) * 1.2

            real_power = FloatData()
            real_power.shorts.s0 = res.registers[7] if not debug else -1
            real_power.shorts.s1 = res.registers[6] if not debug else -1
            data['real_power'] = real_power.float if not debug else randint(1, 100) * 1.2

            reactive_power = FloatData()
            reactive_power.shorts.s0 = res.registers[11] if not debug else -1
            reactive_power.shorts.s1 = res.registers[10] if not debug else -1
            data['reactive_power'] = reactive_power.float if not debug else randint(1, 100) * 1.2

            voltage_LL = FloatData()
            voltage_LL.shorts.s0 = res.registers[15] if not debug else -1
            voltage_LL.shorts.s1 = res.registers[14] if not debug else -1
            data['voltage_LL'] = voltage_LL.float if not debug else randint(1, 100) * 1.2

            frequency = FloatData()
            frequency.shorts.s0 = res.registers[21] if not debug else -1
            frequency.shorts.s1 = res.registers[20] if not debug else -1
            data['frequency'] = frequency.float if not debug else randint(1, 100) * 1.2

            if i == 0:
                self.AC_Meter_0.data.append(data)
            elif i == 1:
                self.AC_Meter_1.data.append(data)

            self.logger.info(data)

    def save_data(self):
        row_begin = 3
        col_begin = 1
        # table header
        self.sheet_for_AC_Meter_0.append(
            ('time', 'real_energy', 'real_power', 'reactive_power', 'voltage_LL', 'frequency'))

        # put the data in
        for row in range(row_begin, len(self.AC_Meter_0.data) + row_begin):
            # for col in range(27, 54):
            self.sheet_for_AC_Meter_0.cell(
                column=col_begin, row=row,
                value="%s" % strftime('%H:%M:%S %d-%b-%Y', localtime(self.AC_Meter_0.data[row - row_begin]['time']))
            )

            self.sheet_for_AC_Meter_0.cell(
                column=col_begin + 1, row=row,
                value="%f" % self.AC_Meter_0.data[row - row_begin]['real_energy']
            ).data_type = 'float'

            self.sheet_for_AC_Meter_0.cell(
                column=col_begin + 2, row=row,
                value="%f" % self.AC_Meter_0.data[row - row_begin]['real_power']
            ).data_type = 'float'
            self.sheet_for_AC_Meter_0.cell(
                column=col_begin + 3, row=row,
                value="%f" % self.AC_Meter_0.data[row - row_begin]['reactive_power']
            ).data_type = 'float'
            self.sheet_for_AC_Meter_0.cell(
                column=col_begin + 4, row=row,
                value="%f" % self.AC_Meter_0.data[row - row_begin]['voltage_LL']
            ).data_type = 'float'
            self.sheet_for_AC_Meter_0.cell(
                column=col_begin + 5, row=row,
                value="%f" % self.AC_Meter_0.data[row - row_begin]['frequency']
            ).data_type = 'float'

        self.sheet_for_AC_Meter_1.append(
            ('time', 'real_energy', 'real_power', 'reactive_power', 'voltage_LL', 'frequency'))
        for row in range(row_begin, len(self.AC_Meter_1.data) + row_begin):
            # for col in range(27, 54):
            self.sheet_for_AC_Meter_1.cell(
                column=col_begin, row=row,
                value="%s" % strftime('%H:%M:%S %d-%b-%Y', localtime(self.AC_Meter_1.data[row - row_begin]['time']))
            )
            self.sheet_for_AC_Meter_1.cell(
                column=col_begin + 1, row=row,
                value="%f" % self.AC_Meter_1.data[row - row_begin]['real_energy']
            ).data_type = 'float'
            self.sheet_for_AC_Meter_1.cell(
                column=col_begin + 2, row=row,
                value="%f" % self.AC_Meter_1.data[row - row_begin]['real_power']
            ).data_type = 'float'
            self.sheet_for_AC_Meter_1.cell(
                column=col_begin + 3, row=row,
                value="%f" % self.AC_Meter_1.data[row - row_begin]['reactive_power']
            ).data_type = 'float'
            self.sheet_for_AC_Meter_1.cell(
                column=col_begin + 4, row=row,
                value="%f" % self.AC_Meter_1.data[row - row_begin]['voltage_LL']
            ).data_type = 'float'
            self.sheet_for_AC_Meter_1.cell(
                column=col_begin + 5, row=row,
                value="%f" % self.AC_Meter_1.data[row - row_begin]['frequency']
            ).data_type = 'float'

        self.book.save(filename=os.path.join(os.path.dirname(os.path.realpath(__file__)),'..','db',self.name +'-'+ strftime('%H.%M.%S %d-%b-%Y') + '.xlsx'))

    def clear_data(self):
        self.AC_Meter_0.data = []
        self.AC_Meter_1.data = []
